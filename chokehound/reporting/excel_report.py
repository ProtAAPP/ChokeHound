"""
Excel report generation module for ChokeHound.

This module handles Excel file creation, formatting, and report generation.
"""

import os
import sys
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from typing import Dict, Optional, List

import chokehound.config.settings as config
from chokehound.config import risk_config


# Mapping of relationship types to their BloodHound documentation URL paths
RELATIONSHIP_TYPE_URL_MAP = {
    # AD / ADCS / Core Edges
    "AbuseTGTDelegation": "abuse-tgt-delegation",
    "ADCSESC1": "adcs-esc1",
    "ADCSESC10a": "adcs-esc10a",
    "ADCSESC10b": "adcs-esc10b",
    "ADCSESC13": "adcs-esc13",
    "ADCSESC3": "adcs-esc3",
    "ADCSESC4": "adcs-esc4",
    "ADCSESC6a": "adcs-esc6a",
    "ADCSESC6b": "adcs-esc6b",
    "ADCSESC9a": "adcs-esc9a",
    "ADCSESC9b": "adcs-esc9b",
    "AddAllowedToAct": "add-allowed-to-act",
    "AddKeyCredentialLink": "add-key-credential-link",
    "AddMember": "add-member",
    "AddSelf": "add-self",
    "AdminTo": "admin-to",
    "AllExtendedRights": "all-extended-rights",
    "AllowedToAct": "allowed-to-act",
    "AllowedToDelegate": "allowed-to-delegate",
    
    # Azure / Entra Edges
    "AZAddMembers": "az-add-members",
    "AZAddOwner": "az-add-owner",
    "AZAddSecret": "az-add-secret",
    "AZAKSContributor": "az-aks-contributor",
    "AZAppAdmin": "az-app-admin",
    "AZAutomationContributor": "az-automation-contributor",
    "AZAvereContributor": "az-avere-contributor",
    "AZCloudAppAdmin": "az-cloud-app-admin",
    "AZContains": "az-contains",
    "AZContributor": "az-contributor",
    "AZExecuteCommand": "az-execute-command",
    "AZGetCertificates": "az-get-certificates",
    "AZGetKeys": "az-get-keys",
    "AZGetSecrets": "az-get-secrets",
    "AZGlobalAdmin": "az-global-admin",
    "AZHasRole": "az-has-role",
    "AZKeyVaultKVContributor": "az-key-vault-contributor",
    "AZLogicAppContributor": "az-logic-app-contributor",
    "AZManagedIdentity": "az-managed-identity",
    "AZMemberOf": "az-member-of",
    "AZMGAddMember": "az-mg-add-member",
    "AZMGAddOwner": "az-mg-add-owner",
    "AZMGAddSecret": "az-mg-add-secret",
    "AZMGAppRoleAssignment_ReadWrite_All": "az-mg-app-role-assignment-readwrite-all",
    "AZMGApplication_ReadWrite_All": "az-mg-application-readwrite-all",
    "AZMGDirectory_ReadWrite_All": "az-mg-directory-readwrite-all",
    "AZMGGrantAppRoles": "az-mg-grant-app-roles",
    "AZMGGrantRole": "az-mg-grant-role",
    "AZMGGroupMember_ReadWrite_All": "az-mg-group-member-readwrite-all",
    "AZMGGroup_ReadWrite_All": "az-mg-group-readwrite-all",
    "AZMGRoleManagement_ReadWrite_Directory": "az-mg-role-management-readwrite-directory",
    "AZMGServicePrincipalEndpoint_ReadWrite_All": "az-mg-service-principal-endpoint-readwrite-all",
    "AZNodeResourceGroup": "az-node-resource-group",
    "AZOwner": "az-owner",
    "AZOwns": "az-owns",
    "AZPrivilegedAuthAdmin": "az-privileged-auth-admin",
    "AZPrivilegedRoleAdmin": "az-privileged-role-admin",
    "AZResetPassword": "az-reset-password",
    "AZRoleApprover": "az-role-approver",
    "AZRoleEligible": "az-role-eligible",
    "AZRunsAs": "az-runs-as",
    "AZScopedTo": "az-scoped-to",
    "AZUserAccessAdministrator": "az-user-access-administrator",
    "AZVMAdminLogin": "az-vm-admin-login",
    "AZVMContributor": "az-vm-contributor",
    "AZWebsiteContributor": "az-website-contributor",
    
    # Lateral movement, coercion, trusts, sessions
    "CanPSRemote": "can-ps-remote",
    "CanRDP": "can-rdp",
    "ClaimSpecialIdentity": "claim-special-identity",
    "CoerceAndRelayNTLMToADCS": "coerce-and-relay-ntlm-to-adcs",
    "CoerceAndRelayNTLMToLDAP": "coerce-and-relay-ntlm-to-ldap",
    "CoerceAndRelayNTLMToLDAPS": "coerce-and-relay-ntlm-to-ldaps",
    "CoerceAndRelayNTLMToSMB": "coerce-and-relay-ntlm-to-smb",
    "CoerceToTGT": "coerce-to-tgt",
    "Contains": "contains",
    "CrossForestTrust": "cross-forest-trust",
    "DCFor": "dc-for",
    "DCSync": "dc-sync",
    "DelegatedEnrollmentAgent": "delegated-enrollment-agent",
    "DumpSMSAPassword": "dump-smsa-password",
    
    # Certificates, GPO, rights, replication
    "Enroll": "enroll",
    "EnrollOnBehalfOf": "enroll-on-behalf-of",
    "EnterpriseCAFor": "enterprise-ca-for",
    "ExecuteDCOM": "execute-dcom",
    "ExtendedByPolicy": "extended-by-policy",
    "ForceChangePassword": "force-change-password",
    "GenericAll": "generic-all",
    "GenericWrite": "generic-write",
    "GetChanges": "get-changes",
    "GetChangesAll": "get-changes-all",
    "GetChangesInFilteredSet": "get-changes-in-filtered-set",
    "GoldenCert": "golden-cert",
    "GPLink": "gp-link",
    
    # Sessions, SID, PKI, ownership
    "HasSession": "has-session",
    "HasSIDHistory": "has-sid-history",
    "HasTrustKeys": "has-trust-keys",
    "HostsCAService": "hosts-ca-service",
    "IssuedSignedBy": "issued-signed-by",
    "LocalToComputer": "local-to-computer",
    "ManageCA": "manage-ca",
    "ManageCertificates": "manage-certificates",
    "MemberOf": "member-of",
    "MemberOfLocalGroup": "member-of-local-group",
    "NTAuthStoreFor": "nt-auth-store-for",
    "OIDGroupLink": "oid-group-link",
    "Owns": "owns",
    "ProtectAdminGroups": "protect-admin-groups",
    "PublishedTo": "published-to",
    
    # Passwords, LAPS, sync, rights abuse
    "ReadGMSAPassword": "read-gmsa-password",
    "ReadLAPSPassword": "read-laps-password",
    "RemoteInteractiveLogonRight": "remote-interactive-logon-right",
    "RootCAFor": "root-ca-for",
    "SameForestTrust": "same-forest-trust",
    "SpoofSIDHistory": "spoof-sid-history",
    "SQLAdmin": "sql-admin",
    "SyncLAPSPassword": "sync-laps-password",
    "SyncedToADUser": "synced-to-ad-user",
    "SyncedToEntraUser": "synced-to-entra-user",
    "TrustedForNTAuth": "trusted-for-nt-auth",
    
    # Write / abuse rights
    "WriteAccountRestrictions": "write-account-restrictions",
    "WriteDacl": "write-dacl",
    "WriteGPLink": "write-gp-link",
    "WriteOwner": "write-owner",
    "WritePKIEnrollmentFlag": "write-pki-enrollment-flag",
    "WritePKINameFlag": "write-pki-name-flag",
    "WriteSPN": "write-spn",
}


class ExcelReportGenerator:
    """Generates Excel reports from security check results."""
    
    def __init__(self, output_filename: str, domains: Optional[List[Dict[str, str]]] = None, tenants: Optional[List[Dict[str, str]]] = None):
        """
        Initialize report generator.
        
        Args:
            output_filename: Output Excel filename
            domains: Optional list of domain dictionaries with 'name' and 'objectid'
            tenants: Optional list of tenant dictionaries with 'name' and 'objectid'
        """
        self.output_filename = output_filename
        self.domains = domains or []
        self.tenants = tenants or []
    
    def detect_limit_in_query(self, query_string: str) -> Optional[int]:
        """
        Detect if a LIMIT clause exists in a Cypher query and extract its value.
        
        Args:
            query_string: Cypher query string
            
        Returns:
            Limit value if found, None otherwise
        """
        # Look for LIMIT followed by a number (case insensitive)
        match = re.search(r'\bLIMIT\s+(\d+)', query_string, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return None
    
    def convert_relationship_type_to_url(self, relationship_type: str) -> str:
        """
        Convert relationship type name to BloodHound documentation URL format.
        
        Args:
            relationship_type: Relationship type string
            
        Returns:
            URL-friendly string (lowercase with hyphens)
        """
        relationship_type = str(relationship_type).strip()
        
        if relationship_type in RELATIONSHIP_TYPE_URL_MAP:
            return RELATIONSHIP_TYPE_URL_MAP[relationship_type]
        
        return relationship_type.lower()
    
    def add_relationship_type_hyperlinks(self, worksheet, df: pd.DataFrame):
        """Add hyperlinks to RelationshipType column cells."""
        if df.empty or 'RelationshipType' not in df.columns:
            return
        
        relationship_type_col = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'RelationshipType':
                relationship_type_col = idx
                break
        
        if relationship_type_col is None:
            return
        
        col_letter = get_column_letter(relationship_type_col)
        base_url = "https://bloodhound.specterops.io/resources/edges/"
        
        for row_idx, relationship_type in enumerate(df['RelationshipType'], start=2):
            if pd.notna(relationship_type) and relationship_type:
                relationship_type_str = str(relationship_type).strip()
                url_suffix = self.convert_relationship_type_to_url(relationship_type_str)
                url = base_url + url_suffix
                
                cell = worksheet[f"{col_letter}{row_idx}"]
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
                cell.value = relationship_type_str
    
    def color_risk_column(self, worksheet, df: pd.DataFrame):
        """Color the RiskScore column based on risk levels."""
        if df.empty or 'RiskScore' not in df.columns:
            return
        
        risk_score_col = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'RiskScore':
                risk_score_col = idx
                break
        
        if risk_score_col is None:
            return
        
        col_letter = get_column_letter(risk_score_col)
        
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row_idx, risk_score in enumerate(df['RiskScore'], start=2):
            if pd.notna(risk_score):
                try:
                    risk_value = float(risk_score)
                    cell = worksheet[f"{col_letter}{row_idx}"]
                    
                    if risk_value >= 60:
                        cell.fill = red_fill
                    elif risk_value >= 31:
                        cell.fill = orange_fill
                    else:
                        cell.fill = yellow_fill
                except (ValueError, TypeError):
                    continue
    
    def format_sheet_as_table(self, worksheet, df: pd.DataFrame):
        """Format a worksheet as an Excel table with filterable/sortable headers."""
        if df.empty:
            return
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row < 2:
            return
        
        start_cell = "A1"
        end_cell = f"{get_column_letter(max_col)}{max_row}"
        table_range = f"{start_cell}:{end_cell}"
        
        table_name = f"Table_{worksheet.title[:20].replace(' ', '_')}"
        table_name = ''.join(c for c in table_name if c.isalnum() or c == '_')
        
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleLight8",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
        
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_cover_sheet(self, workbook):
        """Create a cover sheet for the Excel report."""
        cover_sheet = workbook.create_sheet("Cover", 0)
        cover_sheet.sheet_view.showGridLines = False
        
        logo_paths = [
            os.path.join(os.path.dirname(__file__), "..", "..", "logo.png"),
            os.path.join(os.path.dirname(__file__), "..", "..", "logo.jpg"),
            os.path.join(os.path.dirname(__file__), "..", "..", "logo.jpeg"),
        ]
        
        logo_added = False
        logo_height = 120
        
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                try:
                    img = Image(logo_path)
                    aspect_ratio = img.width / img.height
                    img.height = logo_height
                    img.width = int(logo_height * aspect_ratio)
                    cover_sheet.add_image(img, 'B1')
                    logo_added = True
                    cover_sheet.row_dimensions[1].height = logo_height * 0.75
                    cover_sheet.row_dimensions[2].height = 10
                    cover_sheet.row_dimensions[3].height = 10
                    cover_sheet.row_dimensions[4].height = 10
                    break
                except Exception as e:
                    print(f"[WARNING] Could not add logo from {logo_path}: {e}")
                    continue
        
        title_row = 6 if logo_added else 2
        
        cover_sheet[f'A{title_row}'] = "ChokeHound - Tier 0 Choke Points Analysis Report"
        cover_sheet[f'A{title_row}'].font = Font(size=16, bold=True)
        cover_sheet[f'A{title_row}'].alignment = Alignment(horizontal='center')
        cover_sheet.merge_cells(f'A{title_row}:D{title_row}')
        
        row = title_row + 2
        cover_sheet[f'A{row}'] = "Report Information"
        cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        cover_sheet[f'A{row}'] = "Generated:"
        cover_sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if self.domains:
            row += 1
            cover_sheet[f'A{row}'] = "Active Directory Domains:"
            cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
            for domain in self.domains:
                row += 1
                domain_text = f"  • {domain['name']}"
                if domain.get('objectid'):
                    domain_text += f" (ID: {domain['objectid']})"
                cover_sheet[f'A{row}'] = domain_text
                cover_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                cover_sheet.merge_cells(f'A{row}:D{row}')
        
        if self.tenants:
            row += 1
            cover_sheet[f'A{row}'] = "Azure Tenants:"
            cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
            for tenant in self.tenants:
                row += 1
                tenant_text = f"  • {tenant['name']}"
                if tenant.get('objectid'):
                    tenant_text += f" (ID: {tenant['objectid']})"
                cover_sheet[f'A{row}'] = tenant_text
                cover_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                cover_sheet.merge_cells(f'A{row}:D{row}')
        
        row += 2
        cover_sheet[f'A{row}'] = "Report Description"
        cover_sheet[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        # Build description based on what data exists
        env_text = []
        if self.domains:
            env_text.append("Active Directory")
        if self.tenants:
            env_text.append("Azure/Entra ID")
        
        env_description = " and ".join(env_text) if env_text else "your"
        
        description = (
            f"This report provides a comprehensive security review of {env_description} environment(s) "
            "based on data collected by BloodHound. It identifies various security issues including "
            "choke points, privilege escalation paths, and other security concerns."
        )
        cover_sheet[f'A{row}'] = description
        cover_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        cover_sheet.merge_cells(f'A{row}:D{row + 2}')
        
        cover_sheet.column_dimensions['A'].width = 35
        cover_sheet.column_dimensions['B'].width = 50
        cover_sheet.column_dimensions['C'].width = 15
        cover_sheet.column_dimensions['D'].width = 15
        
        for i in range(1, row + 2):
            cover_sheet.row_dimensions[i].height = 20
    
    def create_summary_sheet(self, workbook, dataframes: Dict[str, pd.DataFrame], query_objects: Dict):
        """Create a summary sheet with query results statistics."""
        summary_sheet = workbook.create_sheet("Summary", 1)
        summary_sheet.sheet_view.showGridLines = False
        
        # Title
        summary_sheet['A1'] = "Query Results Summary"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet['A1'].alignment = Alignment(horizontal='left')
        summary_sheet.merge_cells('A1:D1')
        
        row = 3
        summary_sheet[f'A{row}'] = "This sheet provides an overview of the number of results for each query."
        summary_sheet[f'A{row}'].font = Font(italic=True)
        summary_sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
        summary_sheet.merge_cells(f'A{row}:D{row}')
        row += 2
        
        # Table headers
        headers = ["Query Name", "Result Count", "Limited", "Limit Value"]
        for col_idx, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Add data for each query
        any_limited = False  # Track if any queries are limited
        for sheet_name, df in dataframes.items():
            # Query name with hyperlink
            cell = summary_sheet[f'A{row}']
            cell.value = sheet_name
            safe_sheet_name = sheet_name[:31]
            cell.hyperlink = f"#'{safe_sheet_name}'!A1"
            cell.font = Font(color="0563C1", underline="single")
            
            # Result count
            result_count = len(df)
            summary_sheet[f'B{row}'] = result_count
            summary_sheet[f'B{row}'].alignment = Alignment(horizontal='center')
            
            # Check if results were limited
            is_limited = False
            limit_value = None
            
            if query_objects and sheet_name in query_objects:
                query_obj = query_objects[sheet_name]
                # Get the actual query (handles both static and dynamic queries)
                actual_query = query_obj.get_query() if hasattr(query_obj, 'get_query') else query_obj.cypher_query
                limit_value = self.detect_limit_in_query(actual_query)
                if limit_value is not None and result_count >= limit_value:
                    is_limited = True
                    any_limited = True  # Mark that at least one query is limited
            
            # Limited status
            limited_text = "Yes" if is_limited else "No"
            summary_sheet[f'C{row}'] = limited_text
            summary_sheet[f'C{row}'].alignment = Alignment(horizontal='center')
            if is_limited:
                summary_sheet[f'C{row}'].font = Font(color="FF0000", bold=True)  # Red for limited
            else:
                summary_sheet[f'C{row}'].font = Font(color="008000")  # Green for not limited
            
            # Limit value
            if limit_value is not None:
                summary_sheet[f'D{row}'] = limit_value
                summary_sheet[f'D{row}'].alignment = Alignment(horizontal='center')
            else:
                summary_sheet[f'D{row}'] = "N/A"
                summary_sheet[f'D{row}'].alignment = Alignment(horizontal='center')
                summary_sheet[f'D{row}'].font = Font(italic=True, color="808080")
            
            row += 1
        
        # Add warning note if any queries were limited
        if any_limited:
            row += 2  # Add spacing
            warning_cell = summary_sheet[f'A{row}']
            warning_cell.value = "⚠️ WARNING: Some query results were limited. The limit values can be configured in the settings.py file."
            warning_cell.font = Font(bold=True, color="FF0000", size=11)
            warning_cell.alignment = Alignment(wrap_text=True)
            summary_sheet.merge_cells(f'A{row}:D{row}')
            
            # Add yellow background to warning
            warning_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 50
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 15
    
    def create_documentation_sheet(self, workbook, dataframes: Dict[str, pd.DataFrame], query_descriptions: Dict[str, str], query_objects: Dict, sorted_sheet_names: List[str] = None):
        """Create a documentation sheet explaining what each sheet contains."""
        doc_sheet = workbook.create_sheet("Documentation", 2)
        
        doc_sheet['A1'] = "Report Documentation"
        doc_sheet['A1'].font = Font(size=16, bold=True)
        doc_sheet['A1'].alignment = Alignment(horizontal='left')
        doc_sheet.merge_cells('A1:D1')
        
        row = 3
        doc_sheet[f'A{row}'] = "This document explains the contents of each sheet and the meaning of each column."
        doc_sheet[f'A{row}'].font = Font(italic=True)
        doc_sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
        doc_sheet.merge_cells(f'A{row}:D{row}')
        row += 2
        
        # Use sorted sheet names if provided, otherwise use dataframes keys
        sheet_names_to_iterate = sorted_sheet_names if sorted_sheet_names else list(dataframes.keys())
        
        for sheet_name in sheet_names_to_iterate:
            doc_sheet[f'A{row}'] = f"Sheet: {sheet_name}"
            doc_sheet[f'A{row}'].font = Font(size=14, bold=True)
            doc_sheet[f'A{row}'].alignment = Alignment(horizontal='left')
            row += 1
            
            description = query_descriptions.get(sheet_name, "No description available.")
            doc_sheet[f'A{row}'] = "Description:"
            doc_sheet[f'A{row}'].font = Font(bold=True)
            doc_sheet[f'B{row}'] = description
            doc_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            doc_sheet.merge_cells(f'B{row}:D{row}')
            row += 2
            
            if sheet_name in dataframes:
                actual_columns = list(dataframes[sheet_name].columns)
                if actual_columns:
                    doc_sheet[f'A{row}'] = "Columns:"
                    doc_sheet[f'A{row}'].font = Font(size=12, bold=True)
                    row += 1
                    
                    doc_sheet[f'A{row}'] = ", ".join(actual_columns)
                    doc_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    doc_sheet.merge_cells(f'A{row}:D{row}')
                    row += 1
            
            row += 2
        
        doc_sheet.column_dimensions['A'].width = 25
        doc_sheet.column_dimensions['B'].width = 60
        doc_sheet.column_dimensions['C'].width = 10
        doc_sheet.column_dimensions['D'].width = 10
        
        for i in range(1, row + 2):
            doc_sheet.row_dimensions[i].height = 20
    
    def generate_risk_log(self, risk_breakdowns: List[dict], log_filename: str):
        """Generate a detailed log file explaining risk calculations."""
        try:
            with open(log_filename, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("Risk Calculation Log - ChokeHound - Tier 0 Choke Points Analysis\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Excel Report: {self.output_filename}\n")
                f.write(f"Total Choke Points: {len(risk_breakdowns)}\n\n")
                
                min_score, max_score = risk_config.calculate_risk_score_range()
                f.write("Risk Score Range:\n")
                f.write("-" * 80 + "\n")
                f.write(f"  Theoretical Minimum Risk Score: {min_score:.2f}\n")
                f.write(f"  Theoretical Maximum Risk Score: {max_score:.2f}\n")
                f.write(f"  Normalized Range: 1-100 (where {min_score:.2f} = 1 and {max_score:.2f} = 100)\n\n")
                
                f.write("Risk Calculation Formula:\n")
                f.write("-" * 80 + "\n")
                f.write("Raw Risk Score = (SourceObjectWeight × SourceObjectCategory) +\n")
                f.write("                 (RelationshipTypeWeight × RelationshipTypeCategory) +\n")
                f.write("                 (TargetObjectWeight × TargetObjectCategory) +\n")
                f.write("                 (AffectedAttackPathsWeight × PathsMultiplier × 10)\n\n")
                
                f.write("Normalization Formula:\n")
                f.write("-" * 80 + "\n")
                f.write("Normalized Risk Score (1-100) = 1 + ((RawRiskScore - MinScore) / (MaxScore - MinScore)) × 99\n")
                f.write(f"Where MinScore = {min_score:.2f} and MaxScore = {max_score:.2f}\n\n")
                
                f.write("Component Weights:\n")
                f.write("-" * 80 + "\n")
                weights = risk_config.RISK_WEIGHTS
                f.write(f"  - Source Object Weight: {weights['source_object']}\n")
                f.write(f"  - Relationship Type Weight: {weights['relationship_type']}\n")
                f.write(f"  - Target Object Weight: {weights['target_object']}\n")
                f.write(f"  - Affected Attack Paths Weight: {weights['affected_attack_paths']}\n\n")
                
                f.write("=" * 80 + "\n")
                f.write("DETAILED RISK CALCULATIONS\n")
                f.write("=" * 80 + "\n\n")
                
                for idx, bd in enumerate(risk_breakdowns, 1):
                    f.write(f"Choke Point #{idx}\n")
                    f.write("-" * 80 + "\n")
                    f.write(f"Source: {bd['source_name']} ({bd['source_type']})\n")
                    f.write(f"  Risk Category: {bd['source_risk_category']}\n")
                    f.write(f"  Weight: {bd['source_weight']}\n")
                    f.write(f"  Component: {bd['source_weight']} × {bd['source_risk_category']} = {bd['source_component']}\n\n")
                    
                    f.write(f"Relationship: {bd['relationship_type']}\n")
                    f.write(f"  Risk Category: {bd['relationship_risk_category']}\n")
                    f.write(f"  Weight: {bd['relationship_weight']}\n")
                    f.write(f"  Component: {bd['relationship_weight']} × {bd['relationship_risk_category']} = {bd['relationship_component']}\n\n")
                    
                    f.write(f"Target: {bd['target_name']} ({bd['target_type']})\n")
                    f.write(f"  Risk Category: {bd['target_risk_category']}\n")
                    f.write(f"  Weight: {bd['target_weight']}\n")
                    f.write(f"  Component: {bd['target_weight']} × {bd['target_risk_category']} = {bd['target_component']}\n\n")
                    
                    f.write(f"Affected Attack Paths: {bd['affected_paths']}\n")
                    f.write(f"  Multiplier: {bd['paths_multiplier']}\n")
                    f.write(f"  Weight: {bd['paths_weight']}\n")
                    f.write(f"  Component: {bd['paths_weight']} × {bd['paths_multiplier']} × 10 = {bd['paths_component']}\n\n")
                    
                    f.write(f"TOTAL RISK SCORE: {bd['source_component']} + {bd['relationship_component']} + ")
                    f.write(f"{bd['target_component']} + {bd['paths_component']} = {bd['total_risk_score']}\n")
                    
                    normalized_score = risk_config.normalize_risk_score(bd['total_risk_score'])
                    f.write(f"NORMALIZED RISK SCORE (1-100): {normalized_score}\n\n")
                    f.write("=" * 80 + "\n\n")
            
            print(f"[OK] Risk calculation log generated: {log_filename}")
        except Exception as e:
            print(f"[WARNING] Error generating risk log: {e}")
    
    def generate(self, dataframes: Dict[str, pd.DataFrame], 
                 query_descriptions: Dict[str, str],
                 query_objects: Optional[Dict[str, any]] = None,
                 risk_breakdowns: Optional[List[dict]] = None,
                 enable_logging: bool = False):
        """
        Generate Excel report from dataframes.
        
        Args:
            dataframes: Dictionary mapping sheet names to pandas DataFrames
            query_descriptions: Dictionary mapping query names to descriptions
            query_objects: Optional dictionary mapping query names to SecurityQuery objects
            risk_breakdowns: Optional list of risk breakdown dictionaries for logging
            enable_logging: If True, generate risk calculation log file
        """
        # Check if file exists and is accessible
        if os.path.exists(self.output_filename):
            try:
                test_file = open(self.output_filename, 'r+b')
                test_file.close()
            except PermissionError:
                print(f"[ERROR] Error: The file '{self.output_filename}' is currently open in another application.")
                print("   Please close the file and try again.")
                sys.exit(1)
            except Exception as e:
                print(f"[ERROR] Error accessing file '{self.output_filename}': {e}")
                sys.exit(1)
        
        # Filter dataframes to exclude empty results
        filtered_dataframes = {}
        
        for sheet_name, df in dataframes.items():
            # Check if dataframe has actual results (not just "No results found" or "Error" messages)
            has_results = False
            if not df.empty:
                # Check if it's an error or info message
                if 'Error' in df.columns or 'Info' in df.columns:
                    has_results = False
                else:
                    has_results = True
            
            if has_results:
                filtered_dataframes[sheet_name] = df
        
        # Use filtered dataframes for report generation
        dataframes = filtered_dataframes
        
        # Use natural order of dataframes
        sorted_sheet_names = list(dataframes.keys())
        
        # Create Excel writer
        try:
            writer = pd.ExcelWriter(self.output_filename, engine="openpyxl", mode='w')
        except PermissionError:
            print(f"[ERROR] Error: Cannot write to '{self.output_filename}'. The file may be open in another application.")
            print("   Please close the file and try again.")
            sys.exit(1)
        except Exception as e:
            print(f"[ERROR] Error creating Excel writer: {e}")
            sys.exit(1)
        
        # Write dataframes to Excel in sorted order
        for sheet_name in sorted_sheet_names:
            df = dataframes[sheet_name]
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
        try:
            writer.close()
        except Exception as e:
            print(f"[ERROR] Error closing Excel writer: {e}")
            sys.exit(1)
        
        # Load workbook to add cover sheet and format tables
        print("Creating cover sheet and formatting tables...")
        try:
            workbook = load_workbook(self.output_filename)
            
            print("  Creating cover sheet...")
            self.create_cover_sheet(workbook)
            
            print("  Creating summary sheet...")
            self.create_summary_sheet(workbook, dataframes, query_objects or {})
            
            print("  Creating documentation sheet...")
            self.create_documentation_sheet(workbook, dataframes, query_descriptions, query_objects or {}, sorted_sheet_names)
            
            # Format all data sheets as tables and add hyperlinks (in sorted order)
            print("  Formatting data sheets...")
            for sheet_name in sorted_sheet_names:
                df = dataframes[sheet_name]
                safe_sheet_name = sheet_name[:31]
                if safe_sheet_name in workbook.sheetnames:
                    worksheet = workbook[safe_sheet_name]
                    try:
                        self.format_sheet_as_table(worksheet, df)
                        self.add_relationship_type_hyperlinks(worksheet, df)
                        self.color_risk_column(worksheet, df)
                    except Exception as e:
                        print(f"  [WARNING] Error formatting sheet '{safe_sheet_name}': {e}")
            
            print("  Saving workbook...")
            workbook.save(self.output_filename)
        except PermissionError:
            print(f"[ERROR] Error: Cannot save '{self.output_filename}'. The file may be open in another application.")
            print("   Please close the file and try again.")
            sys.exit(1)
        except Exception as e:
            print(f"[ERROR] Error saving workbook: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
        
        print(f"[OK] Excel report generated: {self.output_filename}")
        
        # Generate risk calculation log if requested
        if enable_logging and risk_breakdowns:
            base_name = os.path.splitext(os.path.basename(self.output_filename))[0]
            log_filename = f"{base_name}_risk_calculation_log.txt"
            if os.path.dirname(self.output_filename):
                log_filename = os.path.join(os.path.dirname(self.output_filename), log_filename)
            
            self.generate_risk_log(risk_breakdowns, log_filename)



